from flask import Flask, request, jsonify, render_template
import openpyxl

app = Flask(__name__)

# def func1(file_path, sentences):
#     workbook = openpyxl.load_workbook(file_path)
#     results = {sentence: {"found": False, "details": []} for sentence in sentences}
    
#     for sheet in workbook.sheetnames:
#         worksheet = workbook[sheet]
        
#         columns = {cell.value: idx for idx, cell in enumerate(next(worksheet.iter_rows(max_row=1)), start=1)}
#         annotator_col = columns.get("Annotator ID")
#         name_col = columns.get("Name")
#         prompt_col = columns.get("Prompt")
#         res1_col = columns.get("Response 1 Rating")
#         res2_col = columns.get("Response 2 Rating")
        
#         for row in worksheet.iter_rows(min_row=2, values_only=True):
#             for cell in row:
#                 if isinstance(cell, str):
#                     for sentence in sentences:
#                         if sentence in cell:
#                             results[sentence]["found"] = True
#                             results[sentence]["details"].append({
#                                 "Annotator ID": row[annotator_col - 1],
#                                 "Name": row[name_col - 1],
#                                 "Response 1 Rating": row[res1_col - 1],
#                                 "Response 2 Rating": row[res2_col - 1]
#                             })
    
#     return results

# @app.route('/')
# def index():
#     return render_template('singleprompt1.html')

# @app.route('/process', methods=['POST'])
# def process():
#     data = request.json
#     sentences = data.get('sentences')
#     file_path = 'C:/Users/OW_USER/gautam/HTML files/finexo-html/July 15  Bat 13 - Seq 40.xlsx'
#     results = func1(file_path, sentences)
#     return jsonify(results)
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/')
def about():
    return render_template('about.html')

@app.route('/singleprompt1')
def single_prompt():
    return render_template('singleprompt1.html')

@app.route('/sheetfind')
def sheetfind():
    return render_template('sheetfind.html')

@app.route('/multiprompt')
def multiprompt():
    return render_template('multiprompt.html')


def func1(file_path, sentences):
    print(file_path)
    workbook = openpyxl.load_workbook(file_path)
    results = {sentence: {"found": False, "details": []} for sentence in sentences}
    
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        
        columns = {cell.value: idx for idx, cell in enumerate(next(worksheet.iter_rows(max_row=1)), start=1)}
        annotator_col = columns.get("Annotator ID")
        name_col = columns.get("Name")
        prompt_col = columns.get("Prompt")
        res1_col = columns.get("Response 1 Rating")
        res2_col = columns.get("Response 2 Rating")
        
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            for cell in row:
                if isinstance(cell, str):
                    for sentence in sentences:
                        if sentence in cell:
                            results[sentence]["found"] = True
                            results[sentence]["details"].append({
                                "Annotator ID": row[annotator_col - 1],
                                "Name": row[name_col - 1],
                                "Response 1 Rating": row[res1_col - 1],
                                "Response 2 Rating": row[res2_col - 1]
                            })
    return results
@app.route('/process', methods=['POST'])
def process():
    data = request.json
    sentences = data.get('sentences')
    file_path = 'C:/Users/harip/Desktop/office/Restart-Project/harirepo/finexo-html/July 15  Bat 13 - Seq 40.xlsx'
    results = func1(file_path, sentences)
    return jsonify(results)

if __name__ == '__main__':
    app.run(debug=True)